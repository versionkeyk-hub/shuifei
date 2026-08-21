"""Build a traceable product catalog from the user's source workbooks.

This script only reads the source directory. It creates a catalog JSON and an
import manifest so every displayed product fact retains its workbook source.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


CATALOG_VERSION = "2026-08-21"


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").strip()


def numeric(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def compact(value: str) -> str:
    value = value.lower()
    value = value.replace("农小蛙", "").replace("锄头猫", "").replace("安欣农", "")
    value = re.sub(r"[\s.·•、()（）【】\[\]\-_]", "", value)
    value = re.sub(r"(提苗型|平衡型|膨果型|高氮型|高磷型|高钾型)$", "", value)
    return value


def identity(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[\s.·•、()（）【】\[\]\-_]", "", value)
    return value


def stable_id(value: str) -> str:
    normalized = identity(value) or value
    digest = hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:10]
    safe = re.sub(r"[^a-z0-9]+", "-", normalized.encode("unicode_escape").decode("ascii")).strip("-")
    return f"prd-{safe[:24] or 'item'}-{digest}"


def split_product_name(name: str) -> tuple[str, str]:
    normalized = clean(name).replace("·", ".")
    if normalized.startswith("锄头猫."):
        return "锄头猫", normalized.split(".", 1)[1]
    if normalized.startswith("安欣农."):
        return "安欣农", normalized.split(".", 1)[1]
    if normalized.startswith("农小蛙."):
        return "农小蛙", normalized.split(".", 1)[1]
    return "农小蛙", name


def canonical_product_name(name: str) -> str:
    """Normalize known product names while leaving unknown names intact.

    Sales workbooks often append carton quantities to the product-name cell.
    The SKU specification keeps that packaging detail; this function makes the
    product card itself stable across workbook layouts.
    """
    value = clean(name).replace("\n", "").replace("m1", "ml")
    normalized = compact(value)
    aliases = {
        "傲叶鱼蛋白": "傲叶",
        "傲土有机矿源黄腐酸": "傲土",
        "傲岚磷酸二氢钾": "傲岚",
        "傲美硅": "傲美",
        "傲靓lpe": "傲靓",
        "有机水溶肥肥鱼蛋白": "有机水溶肥（鱼蛋白）",
        "锄头猫有机水溶肥肥鱼蛋白": "有机水溶肥（鱼蛋白）",
        "锄头猫有机水溶肥鱼蛋白": "有机水溶肥（鱼蛋白）",
        "锄头猫花果多微量元素": "花果多（微量元素）",
        "锄头猫高钙中量元素": "高钙",
    }
    if normalized in aliases:
        return aliases[normalized]

    anchors = (
        "傲生DP",
        "施可收提苗型",
        "施可收平衡型",
        "施可收膨果型",
        "蓓能高氮型",
        "蓓能高磷型",
        "蓓能高钾型",
        "蓓能平衡型",
        "防冻套装",
        "抗寒套",
        "傲叶",
        "傲土",
        "傲岚",
        "傲果",
        "傲生",
        "傲美",
        "傲脉",
        "傲蕾",
        "傲靓",
        "卓艳",
        "均施",
        "晒安心",
        "果大夫",
        "沣硕",
        "洁特",
        "花大夫",
        "黑岩",
        "锄头猫高钙",
        "锄头猫花果多",
        "锄头猫菌剂",
        "锄头猫氨基酸叶面肥",
        "锄头猫氨基酸水溶肥",
    )
    for anchor in anchors:
        if normalized.startswith(identity(anchor)):
            if anchor == "抗寒套":
                return "防冻套装"
            if anchor == "锄头猫高钙":
                return "高钙"
            if anchor == "锄头猫花果多":
                return "花果多（微量元素）"
            if anchor == "锄头猫菌剂":
                return "微生物菌剂"
            if anchor == "锄头猫氨基酸叶面肥":
                return "氨基酸叶面肥"
            if anchor == "锄头猫氨基酸水溶肥":
                return "氨基酸水溶肥"
            return anchor
    return value


def parse_pack_count(specification: str) -> int | None:
    match = re.search(r"(?:[xX×＊*])\s*(\d+)", specification)
    return int(match.group(1)) if match else None


def sha256(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            hasher.update(block)
    return hasher.hexdigest()


class CatalogBuilder:
    def __init__(self, source_root: Path):
        self.source_root = source_root
        self.products: OrderedDict[str, dict[str, Any]] = OrderedDict()
        self.source_files: list[dict[str, str]] = []
        self.unmatched_rows: list[dict[str, Any]] = []

    def product(self, display_name: str, source_ref: dict[str, str]) -> dict[str, Any]:
        brand, source_name = split_product_name(display_name)
        name = canonical_product_name(source_name)
        key = f"{brand}:{identity(name) or identity(display_name)}"
        if key not in self.products:
            self.products[key] = {
                "id": stable_id(f"{brand}-{name}"),
                "name": name,
                "brand": brand,
                "aliases": [display_name] if display_name != name else [],
                "source_type": "own" if brand in {"农小蛙", "锄头猫", "安欣农"} else "market",
                "form": "",
                "usage": "",
                "plain_usage": "",
                "ingredients": {},
                "specifications": [],
                "source_refs": [],
                "needs_verification": False,
            }
        product = self.products[key]
        for alias in (display_name, source_name):
            if alias not in product["aliases"] and alias != product["name"]:
                product["aliases"].append(alias)
        if source_ref not in product["source_refs"]:
            product["source_refs"].append(source_ref)
        return product

    def add_source_file(self, path: Path) -> None:
        self.source_files.append({
            "path": str(path.relative_to(self.source_root)).replace("\\", "/"),
            "sha256": sha256(path),
        })

    def add_specification(
        self,
        product_name: str,
        specification: str,
        unit: str,
        price: float | None,
        source_ref: dict[str, str],
        sku: str = "",
        product_type: str = "",
        price_tier: str = "标准价",
    ) -> None:
        if not product_name or not specification:
            self.unmatched_rows.append({
                "product_name": product_name,
                "specification": specification,
                "source_ref": source_ref,
            })
            return
        product = self.product(product_name, source_ref)
        item = {
            "sku": sku,
            "specification": specification,
            "unit": unit,
            "inner_pack_count": parse_pack_count(specification),
            "price": price,
            "price_tier": price_tier,
            "product_type": product_type,
            "source_ref": source_ref,
        }
        fingerprint = json.dumps(item, ensure_ascii=False, sort_keys=True)
        existing = {json.dumps(entry, ensure_ascii=False, sort_keys=True) for entry in product["specifications"]}
        if fingerprint not in existing:
            product["specifications"].append(item)

    def read_product_info(self, path: Path) -> None:
        self.add_source_file(path)
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            if "功能成分" not in sheet.title:
                continue
            headers = [clean(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                values = [clean(value) for value in row]
                product_name = values[0] if values else ""
                if not product_name:
                    continue
                source_ref = {"file": str(path.relative_to(self.source_root)).replace("\\", "/"), "sheet": sheet.title, "row": str(row_number)}
                product = self.product(product_name, source_ref)
                product["ingredients"] = {
                    headers[index]: values[index]
                    for index in range(1, min(len(headers), len(values)))
                    if values[index]
                }
                product["usage"] = values[9] if len(values) > 9 else product["usage"]
                product["form"] = values[10] if len(values) > 10 else product["form"]
                product["plain_usage"] = values[12] if len(values) > 12 else product["plain_usage"]
                if len(values) > 11 and values[11]:
                    for item in values[11].split("\n"):
                        match = re.match(r"(.+?)/(\d+(?:\.\d+)?)$", item.strip())
                        if match:
                            self.add_specification(product_name, match.group(1), "", float(match.group(2)), source_ref)

    def read_summary_price_book(self, path: Path) -> None:
        self.add_source_file(path)
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            title = sheet.title
            blank_streak = 0
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                values = [clean(value) for value in row]
                if not any(values):
                    blank_streak += 1
                    if blank_streak >= 30:
                        break
                    continue
                blank_streak = 0
                source_ref = {"file": str(path.relative_to(self.source_root)).replace("\\", "/"), "sheet": title, "row": str(row_number)}
                if title in {"农小蛙", "锄头猫"}:
                    self.add_specification(values[1], values[4], values[5], numeric(row[6]) if len(row) > 6 else None, source_ref, sku=values[0], product_type=values[2])
                else:
                    self.add_specification(values[0], values[1], values[2], numeric(row[3]) if len(row) > 3 else None, source_ref)

    def read_multi_column_price_book(self, path: Path) -> None:
        self.add_source_file(path)
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            previous_names: dict[int, str] = {}
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                for start in range(0, len(row), 4):
                    if start + 2 >= len(row):
                        continue
                    name = clean(row[start]) or previous_names.get(start, "")
                    specification = clean(row[start + 1])
                    price = numeric(row[start + 2])
                    if clean(row[start]):
                        previous_names[start] = clean(row[start])
                    if not name or not specification or price is None:
                        continue
                    source_ref = {"file": str(path.relative_to(self.source_root)).replace("\\", "/"), "sheet": sheet.title, "row": str(row_number)}
                    self.add_specification(name, specification, "", price, source_ref)

    def read_tier_price_book(self, path: Path) -> None:
        self.add_source_file(path)
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            previous_names: dict[int, str] = {}
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                for name_column, spec_column, sku_column, price_column, tier in (
                    (0, 1, 2, 3, "家庭园艺"),
                    (0, 4, 5, 6, "常规"),
                    (0, 7, 8, 9, "大户专享"),
                    (11, 12, None, 13, "补充价目"),
                ):
                    explicit_name = clean(row[name_column]) if name_column < len(row) else ""
                    if explicit_name:
                        previous_names[name_column] = explicit_name
                    name = explicit_name or previous_names.get(name_column, "")
                    specification = clean(row[spec_column]) if spec_column < len(row) else ""
                    price = numeric(row[price_column]) if price_column < len(row) else None
                    if not name or not specification or price is None:
                        continue
                    sku = clean(row[sku_column]) if sku_column is not None and sku_column < len(row) else ""
                    source_ref = {"file": str(path.relative_to(self.source_root)).replace("\\", "/"), "sheet": sheet.title, "row": str(row_number)}
                    self.add_specification(name, specification, "", price, source_ref, sku=sku, price_tier=tier)

    def image_manifest(self) -> list[dict[str, str]]:
        records: list[dict[str, str]] = []
        for image in self.source_root.rglob("*"):
            if image.suffix.lower() not in {".jpg", ".jpeg", ".png", ".webp"}:
                continue
            records.append({
                "source_path": str(image.relative_to(self.source_root)).replace("\\", "/"),
                "file_name": image.name,
                "sha256": sha256(image),
                "r2_key": "products/" + sha256(image) + image.suffix.lower(),
            })
        return sorted(records, key=lambda item: item["source_path"])

    def attach_product_images(self, products: list[dict[str, Any]], images: list[dict[str, str]]) -> None:
        special_terms = {
            "高钙": ("高钙", "中量元素"),
            "花果多（微量元素）": ("花果多", "微量元素"),
            "多肽氨基酸": ("多肽氨基酸", "锄头猫氨基酸"),
            "氨基酸叶面肥": ("锄头猫氨基酸",),
            "微生物菌剂": ("微生物菌剂", "锄头猫微生物"),
            "有机水溶肥（鱼蛋白）": ("有机水溶肥", "鱼蛋白"),
        }
        for product in products:
            candidates = [product["name"], *product.get("aliases", [])]
            candidates.extend(special_terms.get(product["name"], ()))
            candidates = [identity(candidate) for candidate in candidates if len(identity(candidate)) >= 2]
            matched = []
            for image in images:
                file_name = identity(image["file_name"])
                if any(candidate in file_name for candidate in candidates):
                    matched.append(image)
            product["image_sources"] = matched

        matched_keys = {image["r2_key"] for product in products for image in product["image_sources"]}
        for image in images:
            image["matched_product_ids"] = [product["id"] for product in products if image in product["image_sources"]]
            image["import_status"] = "pending" if image["r2_key"] in matched_keys else "unmatched"

    def payload(self) -> dict[str, Any]:
        products = list(self.products.values())
        images = self.image_manifest()
        self.attach_product_images(products, images)
        for product in products:
            product["specifications"].sort(key=lambda item: (item["price"] is None, item["price"] or 0, item["specification"]))
            product["source_refs"].sort(key=lambda item: (item["file"], item["sheet"], item["row"]))
            product["needs_verification"] = not bool(product["usage"]) or not bool(product["specifications"])
        return {
            "catalog_version": CATALOG_VERSION,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "products": sorted(products, key=lambda item: (item["brand"], item["name"])),
            "source_files": self.source_files,
            "image_manifest": images,
            "unmatched_rows": self.unmatched_rows,
        }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-root", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    root = args.source_root.resolve()
    builder = CatalogBuilder(root)
    builder.read_product_info(root / "产品信息单页20260720.xlsx")
    builder.read_summary_price_book(root / "销售信息截止20260803" / "产品售价表汇总.xlsx")
    builder.read_multi_column_price_book(root / "销售信息截止20260803" / "农小蛙产品价格表20260803.xlsx")
    builder.read_tier_price_book(root / "销售信息截止20260803" / "sku(1).xlsx")

    payload = builder.payload()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(args.output),
        "products": len(payload["products"]),
        "specifications": sum(len(product["specifications"]) for product in payload["products"]),
        "image_files": len(payload["image_manifest"]),
        "unmatched_rows": len(payload["unmatched_rows"]),
    }, ensure_ascii=False))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
